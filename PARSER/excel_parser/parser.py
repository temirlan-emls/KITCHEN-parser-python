import re
import string
import requests
import os
import math
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill
from datetime import datetime
import time as bot_delay


now = datetime.now()
time = now.strftime("%d/%m/%Y %H:%M").replace('/', '.').replace(':', '.')
site = 'https://proftorg.kz/'
main_folder_name = 'PROFTORG'


def prod_parser(main_url):
    # GETTING ONE PRODUCT PAGE
    main_r = requests.get(main_url)

    main_soup = BeautifulSoup(main_r.content, 'html.parser')
    all_lis = main_soup.find('ul', class_='cs-product-gallery').find_all('li')

    title = main_soup.find(
        'h1', class_='cs-title').find('span').text.replace(',', ' ').replace('.', ' ').replace('"', '').replace('/', '.')
    title = title.split()



    # CREATE FOLDER
    
    folder_title = ''
    for word in title:
        folder_title += ' ' + word
        folder_title = folder_title.strip()
    print(folder_title)
    # try:
    #     os.mkdir(os.path.join(os.getcwd(), folder_title))
    # except:
    #     os.mkdir(os.path.join(os.getcwd(), (folder_title + ' Повтор')))
    # os.chdir(os.path.join(os.getcwd(), folder_title))

    # CREATE EXCEL BOOK, SHEET, TIME
    sheet_title = ''
    for word in title[:3]:
        sheet_title += ' ' + word
        if (len(sheet_title) >= 31):
            sheet_title = sheet_title[:31]
        
    execl_wb = Workbook()
    execl_sheet = execl_wb.create_sheet(title=sheet_title, index=0)
    sheet_row = 3
    thin = Side(border_style="thin", color="000000")

    execl_sheet.merge_cells('B1:G1')
    merged_cell = execl_sheet.cell(row=1, column=1)
    merged_cell.fill = PatternFill("solid", start_color="edff39")
    merged_cell.alignment = Alignment(
    horizontal="center", vertical="center", wrap_text=True)
    merged_cell.border = Border(
    top=thin, left=thin, right=thin, bottom=thin)
    merged_cell.value = folder_title

    execl_sheet[f'B2'].value = 'ФОТО'
    execl_sheet[f'C2'].value = 'НАЗВАНИЕ'
    execl_sheet[f'D2'].value = 'ОПИСАНИЕ'
    execl_sheet[f'E2'].value = 'ПРОФТОРГ'
    execl_sheet[f'F2'].value = 'КИТЧЕН'
    execl_sheet[f'G2'].value = 'МАРЖА'


    for num in range(7): 
        alphabet_string = string.ascii_uppercase
        alphabet_list = list(alphabet_string)[:7]
        execl_sheet[f'{alphabet_list[num]}2'].alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
        execl_sheet[f'{alphabet_list[num]}2'].border = Border(
        top=thin, left=thin, right=thin, bottom=thin)
        execl_sheet[f'{alphabet_list[num]}2'].fill = PatternFill("solid", start_color="edff39")

    for li in all_lis:
        # GETTING EXACT PRODUCT PAGE
        url = f"{site}{li.find('a', class_='cs-product-gallery__image-link')['href']}"
        r = requests.get(url)
        soup = BeautifulSoup(r.content, 'html.parser')

        # PRODUCT NAME
        prod_name = soup.find('span', attrs={'data-qaid': 'product_name'}).text

        # PRODUCT CODE
        prod_code = soup.find('span', attrs={'data-qaid': 'product_code'})
        if prod_code is None:
            prod_code = soup.find(
                'span', attrs={'data-qaid': 'product_name'}).text
        else:
            prod_code = soup.find(
                'span', attrs={'data-qaid': 'product_code'}).text

        # PRODUCT DESCRIPTION
        stats = ''

        try:
            prod_stat = soup.find(
            'table', class_="b-product-info").text
            stats = prod_stat
        except:
            prod_desk = soup.find(
                'div', attrs={'data-qaid': 'product_description'})
            if prod_desk is None:
                try:
                    stats = soup.find(
                        'div', attrs={'data-qaid': 'product_description'}).text
                except:
                    stats = 'PUSTO'
            elif prod_desk.find('ul') is None:
                try:
                    stats = soup.find(
                        'div', attrs={'data-qaid': 'product_description'}).text
                except:
                    stats = 'PUSTO'
            else:
                for li in prod_desk.find('ul').find_all('li'):
                    stats = stats + ' ' + li.text
                    
        
        # prod_desk = soup.find(
        #     'div', attrs={'data-qaid': 'product_description'})
        # if prod_desk is None:
        #     try:
        #         stats = soup.find(
        #             'div', attrs={'data-qaid': 'product_description'}).text
        #     except:
        #         stats = 'PUSTO'
        # elif prod_desk.find('ul') is None:
        #     try:
        #         stats = soup.find(
        #         'div', attrs={'data-qaid': 'product_description'}).text
        #     except:
        #         stats = 'PUSTO'
        # else:
        #     for li in prod_desk.find('ul').find_all('li'):
        #         stats = stats + ' ' + li.text


        # PRICE

        try:
            price_tmp = ''

            prod_price_str = soup.find(
            'span', attrs={'data-qaid': 'product_price'}).text
            prod_price_list = re.findall(r'\d+', prod_price_str)
            for num in prod_price_list:
                price_tmp +=  str(num)

            prod_price = int(price_tmp)
            prof_price = prod_price
            kitchen_price = math.ceil((prof_price)*1.25)
            margin = kitchen_price - prof_price
        except:
            pass

        # PRODUCT PHOTO

        try:
            photo_url = soup.find(
                'img', class_="cs-product-image__img csjs-image")['src']
        except:
            pass

        def data_to_sheet(ch, cw):
            try:
                execl_sheet[f'A{sheet_row}'].hyperlink = photo_url
                execl_sheet[f'A{sheet_row}'].style = "Hyperlink"
            except:
                execl_sheet[f'A{sheet_row}'].value = prod_code

            execl_sheet[f'A{sheet_row}'].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            execl_sheet[f'A{sheet_row}'].border = Border(
                top=thin, left=thin, right=thin, bottom=thin)
            execl_sheet.row_dimensions[sheet_row].height = ch
            execl_sheet.column_dimensions['A'].width = cw*1.2

            execl_sheet[f'B{sheet_row}'].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            execl_sheet[f'B{sheet_row}'].border = Border(
                top=thin, left=thin, right=thin, bottom=thin)
            execl_sheet.row_dimensions[sheet_row].height = ch
            execl_sheet.column_dimensions['B'].width = cw*1.2

            execl_sheet[f'C{sheet_row}'].value = prod_code

            execl_sheet[f'C{sheet_row}'].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            execl_sheet[f'C{sheet_row}'].border = Border(
                top=thin, left=thin, right=thin, bottom=thin)
            execl_sheet.row_dimensions[sheet_row].height = ch
            execl_sheet.column_dimensions['C'].width = cw*1.5

            execl_sheet[f'D{sheet_row}'].value = stats

            execl_sheet[f'D{sheet_row}'].alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            execl_sheet[f'D{sheet_row}'].border = Border(
                top=thin, left=thin, right=thin, bottom=thin)
            execl_sheet.row_dimensions[sheet_row].height = ch
            execl_sheet.column_dimensions['D'].width = cw*1.5


            try:
                execl_sheet[f'E{sheet_row}'].value = prof_price
                execl_sheet[f'F{sheet_row}'].value = kitchen_price
                execl_sheet[f'G{sheet_row}'].value = margin

                for num in range(3): 
                    alphabet_string = string.ascii_uppercase
                    alphabet_list = list(alphabet_string)[4:7]
                    execl_sheet[f'{alphabet_list[num]}{sheet_row}'].alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
                    execl_sheet[f'{alphabet_list[num]}{sheet_row}'].border = Border(
                    top=thin, left=thin, right=thin, bottom=thin)

            except:
                execl_sheet[f'E{sheet_row}'].value = 'NO PRICE'
                execl_sheet[f'F{sheet_row}'].value = 'NO PRICE'
                execl_sheet[f'G{sheet_row}'].value = 'NO PRICE'

                for num in range(3): 
                    alphabet_string = string.ascii_uppercase
                    alphabet_list = list(alphabet_string)[4:7]
                    execl_sheet[f'{alphabet_list[num]}{sheet_row}'].alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True)
                    execl_sheet[f'{alphabet_list[num]}{sheet_row}'].border = Border(
                    top=thin, left=thin, right=thin, bottom=thin)



        
        data_to_sheet(100, 20)
        sheet_row += 1

        print(f'{sheet_row-3}/{len(all_lis)} ЗАПИСЬ: {prod_name} {prod_code} ')

    execl_wb.save(filename=(f'{folder_title.strip()}.xlsx'))
    
    # os.chdir("..")


prof_r = requests.get(f"{site}/")
prof_soup = BeautifulSoup(prof_r.content, 'html.parser')
prof_lis = prof_soup.find(
    'ul', class_='cs-product-groups-gallery').find_all('li')

try:
    os.mkdir(os.path.join(os.getcwd(), f'{main_folder_name} {time}'))
except:
    pass
os.chdir(os.path.join(os.getcwd(), f'{main_folder_name} {time}'))

tic = bot_delay.perf_counter()

incre = 1
for li in prof_lis:
    bot_delay.sleep(1.5)
    prod_url = f"{li.find('a', class_='cs-product-groups-gallery__title')['href']}"
    print(f'{incre}/{len(prof_lis)}')
    incre += 1
    prod_parser(f"{site}{prod_url}?product_items_per_page=48")

toc = bot_delay.perf_counter()

secs = toc - tic

print(f"DONE! Time: {secs/60} mins  {secs} secs")